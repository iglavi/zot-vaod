"""
סקריפר HTML — מוריד מטאדאטא מנט המשפט ישירות דרך HTTP (ללא דפדפן).
מהיר פי ~8 מהסקריפר הנוכחי, ומחזיר גם CaseID / DocumentID / IsIDCPublished.

הרצה:
    python scraper_html.py config_html.json

קובץ קונפיג לדוגמה:
{
  "output_dir":      "D:\\Research-bot\\2020",
  "progress_file":   "D:\\Research-bot\\2020\\progress_html.json",
  "log_file":        "D:\\Research-bot\\2020\\scraper_html_log.txt",
  "date_from":       "2020-01-01",
  "date_to":         "2020-12-31",
  "decision_types":  ["פסק דין", "גזר דין", "הכרעת דין"]
}

פלט: metadata.csv + metadata.xlsx באותה תיקייה.
"""

from __future__ import annotations

import csv
import html as html_lib
import json
import sys
import time
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from urllib.parse import urljoin

try:
    import httpx
    from bs4 import BeautifulSoup
except ImportError:
    print("חסרות תלויות. הרץ: pip install httpx beautifulsoup4 openpyxl")
    raise SystemExit(1)

# ── קונפיג ──────────────────────────────────────────────────
_config_path = sys.argv[1] if len(sys.argv) > 1 else "config_html.json"
_cfg = json.loads(Path(_config_path).read_text(encoding="utf-8"))

OUTPUT_DIR     = Path(_cfg["output_dir"])
PROGRESS_FILE  = Path(_cfg.get("progress_file", OUTPUT_DIR / "progress_html.json"))
LOG_FILE       = Path(_cfg.get("log_file",      OUTPUT_DIR / "scraper_html_log.txt"))
DATE_FROM      = datetime.strptime(_cfg["date_from"], "%Y-%m-%d").date()
DATE_TO        = datetime.strptime(_cfg["date_to"],   "%Y-%m-%d").date()
DECISION_TYPES = _cfg.get("decision_types", ["פסק דין", "גזר דין", "הכרעת דין"])
CSV_FILE       = OUTPUT_DIR / "metadata.csv"

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
PROGRESS_FILE.parent.mkdir(parents=True, exist_ok=True)

# ── URL-ים ──────────────────────────────────────────────────
BASE_URL   = "https://www.court.gov.il/NGCS.Web.Site"
HOME_URL   = f"{BASE_URL}/HomePage.aspx"
SEARCH_URL = f"{BASE_URL}/LocateDecisions/LocateDecisionQuering.aspx"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "he-IL,he;q=0.9,en-US;q=0.8,en;q=0.7",
}

# עמודות הפלט — מה מוכר + חדשות
CSV_COLUMNS = [
    "תאריך מתן החלטה",
    "בית משפט",
    "מספר תיק",
    "שם תיק",
    "סוג החלטה",
    "CaseID",
    "DocumentID",
    "DecisionID",
    "IsIDCPublished",
]

# ── לוג ─────────────────────────────────────────────────────
_log_handle = LOG_FILE.open("a", encoding="utf-8")

def log(msg: str):
    ts = datetime.now().strftime("%H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line)
    _log_handle.write(line + "\n")
    _log_handle.flush()


# ── פרוגרס ──────────────────────────────────────────────────
def load_progress() -> set[str]:
    if PROGRESS_FILE.exists() and PROGRESS_FILE.stat().st_size > 0:
        try:
            data = json.loads(PROGRESS_FILE.read_text(encoding="utf-8"))
            return set(data.get("done", []))
        except Exception:
            pass
    return set()

def save_progress(done: set[str]):
    tmp = PROGRESS_FILE.with_suffix(".tmp")
    tmp.write_text(
        json.dumps({"done": sorted(done)}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    tmp.replace(PROGRESS_FILE)


# ── HTTP helpers ─────────────────────────────────────────────
def decode_text(response: httpx.Response) -> str:
    enc = response.encoding or "windows-1255"
    try:
        return response.content.decode(enc)
    except UnicodeDecodeError:
        return response.content.decode("windows-1255", errors="replace")

def parse_form_fields(html: str) -> dict[str, str]:
    soup = BeautifulSoup(html, "html.parser")
    form = soup.find("form", {"id": "Form1"}) or soup.find("form", {"name": "Form1"})
    if form is None:
        raise RuntimeError("לא נמצא Form1")
    fields: dict[str, str] = {}
    for el in form.find_all(["input", "select", "textarea"]):
        name = el.get("name")
        if not name:
            continue
        if el.name == "input":
            t = (el.get("type") or "text").lower()
            if t in {"submit", "button", "image", "file"}:
                continue
            if t in {"radio", "checkbox"}:
                if el.has_attr("checked"):
                    fields[name] = el.get("value", "on")
                continue
            fields[name] = el.get("value", "")
        elif el.name == "select":
            opt = el.find("option", selected=True) or el.find("option")
            if opt:
                fields[name] = opt.get("value", "")
        elif el.name == "textarea":
            fields[name] = el.get_text()
    return fields

def get_dropdown_options(html: str, select_name: str) -> dict[str, str]:
    """מחזיר {label: value} מ-dropdown לפי name."""
    soup = BeautifulSoup(html, "html.parser")
    sel = soup.find("select", {"name": select_name})
    if not sel:
        return {}
    return {opt.get_text(strip=True): opt.get("value", "") for opt in sel.find_all("option")}

def parse_results(html: str) -> list[dict]:
    soup = BeautifulSoup(html, "html.parser")
    store = soup.find("input", {"id": "LocateDecisionsGridArrayStore"})
    if not store or not store.get("value"):
        return []
    return json.loads(html_lib.unescape(store["value"]))


# ── Session ──────────────────────────────────────────────────
class NgcsSession:
    def __init__(self):
        self.client = httpx.Client(headers=HEADERS, follow_redirects=True, timeout=60)
        self.court_map:    dict[str, str] = {}   # שם_ערכאה -> court_id
        self.dec_type_map: dict[str, str] = {}   # שם_סוג -> dec_type_id
        self._search_form_html: str = ""

    def close(self):
        self.client.close()

    def init(self):
        self._reload_search_form()
        self.court_map    = get_dropdown_options(
            self._search_form_html, "LocateByParameters1:ddlSelectCourt")
        self.dec_type_map = get_dropdown_options(
            self._search_form_html, "LocateByParameters1:ddlDecisionType")
        log(f"נטענו {len(self.court_map)} ערכאות, {len(self.dec_type_map)} סוגי החלטה")

    def _reload_search_form(self):
        r = self.client.get(HOME_URL)
        r.raise_for_status()
        fields = parse_form_fields(decode_text(r))
        fields["__EVENTTARGET"]   = "Header1$UpperMenu1$btnVerdictLocalization"
        fields["__EVENTARGUMENT"] = ""
        r = self.client.post(HOME_URL, data=fields)
        r.raise_for_status()
        self._search_form_html = decode_text(r)

    def search(
        self,
        date_str: str,            # DD/MM/YYYY
        dec_type_id: str = "-1",  # -1 = כל סוגי ההחלטות
        court_id:    str = "-1",  # -1 = כל הערכאות
    ) -> list[dict]:
        self._reload_search_form()
        fields = parse_form_fields(self._search_form_html)
        fields["hdnSelectedTab"]                              = "1"
        fields["LocateByParameters1:ddlSelectCourt"]         = court_id
        fields["LocateByParameters1:ddlSelectProceeding"]    = "-1"
        fields["LocateByParameters1:ddlSelectCaseType"]      = "-1"
        fields["LocateByParameters1:ddlSelectCaseInterest"]  = "-1"
        fields["LocateByParameters1:ddlDecisionType"]        = dec_type_id
        fields["LocateByParameters1:dateFrom"]               = date_str
        fields["LocateByParameters1:DateTo"]                 = date_str
        fields["__EVENTTARGET"]                              = "ButtonsGroup1$btnLocate"
        fields["__EVENTARGUMENT"]                            = ""
        r = self.client.post(SEARCH_URL, data=fields)
        r.raise_for_status()
        return parse_results(decode_text(r))


# ── CSV ──────────────────────────────────────────────────────
def open_csv_writer():
    file_exists = CSV_FILE.exists() and CSV_FILE.stat().st_size > 0
    fh = CSV_FILE.open("a", newline="", encoding="utf-8-sig")
    writer = csv.DictWriter(fh, fieldnames=CSV_COLUMNS)
    if not file_exists:
        writer.writeheader()
    return fh, writer

def load_existing_keys() -> set[str]:
    """מספרי תיק שכבר קיימים ב-CSV — למניעת כפילויות."""
    if not CSV_FILE.exists():
        return set()
    seen = set()
    with CSV_FILE.open(encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            key = f"{row.get('תאריך מתן החלטה','')}|{row.get('מספר תיק','')}|{row.get('DocumentID','')}"
            seen.add(key)
    return seen

def result_to_row(res: dict, dec_type_name: str) -> dict:
    date_raw = str(res.get("DecisionSignatureDate", ""))
    # פורמט שהאתר מחזיר: "/Date(1234567890000)/" או "DD/MM/YYYY HH:MM:SS"
    if date_raw.startswith("/Date("):
        ts = int(date_raw[6:date_raw.index(")")])
        dt = datetime.utcfromtimestamp(ts / 1000)
        date_fmt = dt.strftime("%d/%m/%Y %H:%M:%S")
    else:
        date_fmt = date_raw

    return {
        "תאריך מתן החלטה": date_fmt,
        "בית משפט":        str(res.get("CourtName", "")).strip(),
        "מספר תיק":        str(res.get("CaseDisplayIdentifier", "")).strip(),
        "שם תיק":          str(res.get("CaseName", "")).strip(),
        "סוג החלטה":       dec_type_name,
        "CaseID":          res.get("CaseID", ""),
        "DocumentID":      res.get("DocumentID", ""),
        "DecisionID":      res.get("DecisionID", ""),
        "IsIDCPublished":  res.get("IsIDCPublished", ""),
    }

def csv_row_key(row: dict) -> str:
    return f"{row['תאריך מתן החלטה']}|{row['מספר תיק']}|{row['DocumentID']}"


# ── Excel ────────────────────────────────────────────────────
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        log("openpyxl לא מותקן — מדלג על ייצוא Excel (pip install openpyxl)")
        return

    xlsx_path = CSV_FILE.with_suffix(".xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "מטאדאטא"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell_font   = Font(name="Arial", size=9)

    with CSV_FILE.open(encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        for r_idx, row in enumerate(reader, start=1):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font = header_font if r_idx == 1 else cell_font
                cell.alignment = Alignment(horizontal="right", wrap_text=False)
                if r_idx == 1:
                    cell.fill = header_fill

    # רוחב עמודות
    col_widths = {
        "תאריך מתן החלטה": 22, "בית משפט": 30, "מספר תיק": 24,
        "שם תיק": 40, "סוג החלטה": 16, "CaseID": 12,
        "DocumentID": 12, "DecisionID": 12, "IsIDCPublished": 15,
    }
    for i, col in enumerate(CSV_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = col_widths.get(col, 14)

    ws.freeze_panes = "A2"
    ws.sheet_view.rightToLeft = True
    wb.save(xlsx_path)
    log(f"Excel נשמר: {xlsx_path}")


# ── לוגיקה ראשית ─────────────────────────────────────────────
def iter_days(from_date: date, to_date: date):
    d = from_date
    while d <= to_date:
        yield d
        d += timedelta(days=1)

def with_retry(fn, attempts=3, delay=5):
    for attempt in range(attempts):
        try:
            return fn()
        except Exception as e:
            if attempt == attempts - 1:
                raise
            log(f"  [retry {attempt+1}/{attempts}] {e} — ממתין {delay}שנ'...")
            time.sleep(delay)
            delay *= 2


def main():
    log("=" * 60)
    log(f"scraper_html.py — {DATE_FROM} עד {DATE_TO}")

    done      = load_progress()
    seen_keys = load_existing_keys()
    log(f"שילובים שכבר בוצעו: {len(done)}  |  שורות קיימות ב-CSV: {len(seen_keys)}")

    sess = NgcsSession()
    sess.init()

    # בניית מיפוי הפוך: שם -> ID לסוגי החלטה
    # נחפש לפי שם מדויק (כפי שהאתר מחזיר)
    def find_dec_id(name: str) -> str | None:
        # ניסיון 1: שם מדויק
        if name in sess.dec_type_map:
            return sess.dec_type_map[name]
        # ניסיון 2: חיפוש חלקי
        for k, v in sess.dec_type_map.items():
            if name in k or k in name:
                return v
        return None

    stats = {"added": 0, "skipped_dup": 0, "overflow": 0}

    csv_fh, csv_writer = open_csv_writer()

    try:
        for day in iter_days(DATE_FROM, DATE_TO):
            date_str = day.strftime("%d/%m/%Y")   # DD/MM/YYYY לאתר
            date_key = day.strftime("%Y-%m-%d")   # YYYY-MM-DD לפרוגרס

            for dt_name in DECISION_TYPES:
                prog_key = f"{date_key}|{dt_name}"
                if prog_key in done:
                    continue

                dt_id = find_dec_id(dt_name)
                if dt_id is None:
                    log(f"  ⚠ לא נמצא ID לסוג '{dt_name}' — מדלג")
                    continue

                # חיפוש ראשוני: כל הערכאות
                try:
                    results = with_retry(lambda: sess.search(date_str, dt_id, "-1"))
                except Exception as e:
                    log(f"  ✗ שגיאה [{date_str} {dt_name}]: {e}")
                    continue

                if len(results) == 0:
                    done.add(prog_key)
                    save_progress(done)
                    continue

                if len(results) < 100:
                    # אין גלישה — שומרים הכל
                    rows_to_add = [result_to_row(r, dt_name) for r in results]
                    added = 0
                    for row in rows_to_add:
                        k = csv_row_key(row)
                        if k not in seen_keys:
                            csv_writer.writerow(row)
                            seen_keys.add(k)
                            added += 1
                    stats["added"] += added
                    stats["skipped_dup"] += len(rows_to_add) - added
                    log(f"  [{date_str}] [{dt_name}] — {len(results)} תוצאות, נוספו {added}")
                    done.add(prog_key)
                    save_progress(done)

                else:
                    # גלישה — חלק לפי ערכאה
                    log(f"  [{date_str}] [{dt_name}] — 100 תוצאות, מפצל לפי ערכאה...")

                    all_court_done = True
                    for court_name, court_id in sorted(sess.court_map.items()):
                        if court_id in {"-1", "0", ""}:
                            continue
                        court_prog = f"{date_key}|{dt_name}|{court_id}"
                        if court_prog in done:
                            continue

                        try:
                            c_results = with_retry(
                                lambda cid=court_id: sess.search(date_str, dt_id, cid)
                            )
                        except Exception as e:
                            log(f"    ✗ שגיאה ב-{court_name}: {e}")
                            all_court_done = False
                            continue

                        if len(c_results) == 100:
                            log(f"    ⚠ {court_name} — 100 תוצאות (מוגבל!)")
                            stats["overflow"] += 1

                        added = 0
                        for r in c_results:
                            row = result_to_row(r, dt_name)
                            k = csv_row_key(row)
                            if k not in seen_keys:
                                csv_writer.writerow(row)
                                seen_keys.add(k)
                                added += 1
                        if c_results:
                            log(f"    {court_name} — {len(c_results)} תוצאות, נוספו {added}")
                        stats["added"] += added
                        done.add(court_prog)
                        save_progress(done)
                        time.sleep(0.3)

                    if all_court_done:
                        done.add(prog_key)
                        save_progress(done)

                csv_fh.flush()
                time.sleep(0.3)

    finally:
        csv_fh.close()

    log("\n" + "=" * 60)
    log(f"סיכום: נוספו {stats['added']} שורות | "
        f"כפילויות {stats['skipped_dup']} | "
        f"גלישות ב-100 {stats['overflow']}")
    log("=" * 60)

    log("מייצא Excel...")
    export_excel()

    sess.close()
    _log_handle.close()


if __name__ == "__main__":
    main()
